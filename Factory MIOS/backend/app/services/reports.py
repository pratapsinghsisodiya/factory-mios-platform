"""Build .xlsx reports from telemetry. Pulls JSON-shaped telemetry from the DB,
aggregates it, and writes a styled multi-sheet workbook with openpyxl."""
import io
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from sqlalchemy import func
from sqlalchemy.orm import Session
from app.models.models import Device, Telemetry

HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0EA5E9")
TITLE = Font(bold=True, size=14, color="0F172A")
THIN = Border(*[Side(style="thin", color="E2E8F0")] * 4)


def _last(db, tenant_id, device_id, param, since):
    row = (db.query(Telemetry).filter(
        Telemetry.tenant_id == tenant_id, Telemetry.device_id == device_id,
        Telemetry.parameter == param, Telemetry.ts >= since)
        .order_by(Telemetry.ts.desc()).first())
    return row.value if row and row.value is not None else None


def _agg(db, tenant_id, device_id, param, since, fn):
    return db.query(fn(Telemetry.value)).filter(
        Telemetry.tenant_id == tenant_id, Telemetry.device_id == device_id,
        Telemetry.parameter == param, Telemetry.ts >= since).scalar()


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN


def build_report(db: Session, tenant_id: str, report_type: str,
                 window_minutes: int = 1440) -> bytes:
    since = datetime.now(timezone.utc) - timedelta(minutes=window_minutes)
    devices = db.query(Device).filter(Device.tenant_id == tenant_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Factory MIOS — {report_type.upper()} Report"
    ws["A1"].font = TITLE
    ws["A2"] = f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC} · window: last {window_minutes} min"
    ws["A2"].font = Font(italic=True, color="64748B")

    headers = ["Device", "Machine Type", "Availability %", "Performance %",
               "Quality %", "OEE %", "Good", "Reject", "Avg Cycle (s)",
               "Energy (kWh)", "Downtime (min)"]
    hr = 4
    ws.append([])  # spacer to row 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, len(headers))

    r = hr + 1
    for dev in devices:
        good = _last(db, tenant_id, dev.id, "good_count", since) or 0
        total = _last(db, tenant_id, dev.id, "total_count", since) or 0
        reject = _last(db, tenant_id, dev.id, "reject_count", since) or max(total - good, 0)
        downtime = _last(db, tenant_id, dev.id, "downtime_min", since) or 0
        avg_cycle = _agg(db, tenant_id, dev.id, "cycle_time", since, func.avg) or 0
        run_avg = _agg(db, tenant_id, dev.id, "running", since, func.avg)
        energy = _agg(db, tenant_id, dev.id, "energy_kw", since, func.avg) or 0

        availability = round((run_avg or 0) * 100, 1)
        quality = round((good / total * 100) if total else 0, 1)
        # performance vs 30s target cycle, capped at 100
        performance = round(min(30.0 / avg_cycle, 1.0) * 100, 1) if avg_cycle else 0
        oee = round(availability / 100 * performance / 100 * quality / 100 * 100, 1)
        energy_kwh = round(energy * window_minutes / 60, 2)

        ws.cell(row=r, column=1, value=dev.name)
        ws.cell(row=r, column=2, value=dev.machine_type or "—")
        for col, val in enumerate([availability, performance, quality, oee, int(good),
                                   int(reject), round(avg_cycle, 1), energy_kwh,
                                   round(downtime, 1)], start=3):
            ws.cell(row=r, column=col, value=val)
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = THIN
        r += 1

    for col, width in zip("ABCDEFGHIJK", [18, 16, 14, 14, 12, 10, 8, 8, 12, 12, 13]):
        ws.column_dimensions[col].width = width

    # OEE bar chart if we have data
    if r > hr + 1:
        chart = BarChart()
        chart.title = "OEE % by device"
        chart.y_axis.title = "OEE %"
        data = Reference(ws, min_col=6, min_row=hr, max_row=r - 1)
        cats = Reference(ws, min_col=1, min_row=hr + 1, max_row=r - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height, chart.width = 8, 18
        ws.add_chart(chart, f"A{r + 2}")

    # Detail sheet: recent raw telemetry (latest 500 rows)
    ws2 = wb.create_sheet("Telemetry")
    for i, h in enumerate(["Timestamp", "Device", "Parameter", "Value"], start=1):
        ws2.cell(row=1, column=i, value=h)
    _style_header(ws2, 1, 4)
    dev_name = {d.id: d.name for d in devices}
    rows = (db.query(Telemetry).filter(Telemetry.tenant_id == tenant_id,
            Telemetry.ts >= since).order_by(Telemetry.ts.desc()).limit(500).all())
    for idx, t in enumerate(rows, start=2):
        ws2.cell(row=idx, column=1, value=t.ts.strftime("%Y-%m-%d %H:%M:%S"))
        ws2.cell(row=idx, column=2, value=dev_name.get(t.device_id, t.device_id))
        ws2.cell(row=idx, column=3, value=t.parameter)
        ws2.cell(row=idx, column=4, value=t.value)
    for col, width in zip("ABCD", [20, 18, 16, 12]):
        ws2.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
